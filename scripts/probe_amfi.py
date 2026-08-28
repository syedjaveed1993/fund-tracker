"""
Probe #2: AMFI monthly portfolio disclosures.

Goal: find out (a) whether the disclosure page is reachable and
parseable, (b) what the download links look like, and (c) what is
actually inside one of the Excel workbooks.

Runs on GitHub Actions. Read the log, then delete this file.
"""

import io
import re
import ssl
import sys
import urllib.request
import urllib.error

PAGES = [
    "https://www.amfiindia.com/online-center/portfolio-disclosure",
    "https://www.amfiindia.com/otherdata/scheme-wise-disclosure",
    "https://www.amfiindia.com/research-information/amfi-monthly",
]

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/122.0 Safari/537.36")


def get(url, timeout=45):
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE          # AMFI's chain is sometimes incomplete
    req = urllib.request.Request(url, headers={
        "User-Agent": UA,
        "Accept": "text/html,application/xhtml+xml,application/vnd.ms-excel,*/*",
        "Accept-Language": "en-US,en;q=0.9",
    })
    with urllib.request.urlopen(req, timeout=timeout, context=ctx) as r:
        return r.status, r.headers.get("Content-Type", "?"), r.read()


def main():
    print("=" * 72)
    print("STEP 1 - CAN WE REACH THE AMFI PAGES?")
    print("=" * 72)

    html_pages = {}
    for url in PAGES:
        try:
            status, ctype, body = get(url)
            txt = body.decode("utf-8", "replace")
            html_pages[url] = txt
            print("  OK   %s  [%s]  %d bytes" % (url, ctype, len(body)))
        except urllib.error.HTTPError as e:
            print("  %-4s %s" % (e.code, url))
        except Exception as e:
            print("  ERR  %s  (%s: %s)" % (url, type(e).__name__, e))

    if not html_pages:
        print("\n  Could not reach any AMFI page. Stopping.")
        return

    print()
    print("=" * 72)
    print("STEP 2 - WHAT DOWNLOAD LINKS ARE ON THOSE PAGES?")
    print("=" * 72)

    candidates = []
    for url, txt in html_pages.items():
        links = re.findall(r'href=["\']([^"\']+)["\']', txt, re.I)
        data_links = [l for l in links
                      if re.search(r"\.(xls|xlsx|csv|zip)(\?|$)", l, re.I)
                      or "portfolio" in l.lower() or "disclosure" in l.lower()]
        print("\n  --- %s" % url)
        print("      %d links total, %d look like data/disclosure links" % (len(links), len(data_links)))
        for l in data_links[:30]:
            full = l if l.startswith("http") else ("https://www.amfiindia.com" + ("" if l.startswith("/") else "/") + l)
            print("       ", full[:150])
            if re.search(r"\.(xls|xlsx|csv)(\?|$)", l, re.I):
                candidates.append(full)

        # AMFI often builds these lists with JS from an embedded blob
        for kw in ["portfolioDisclosure", "PortfolioData", "amfi-portfolio", "fileUrl", "DownloadUrl"]:
            if kw in txt:
                i = txt.find(kw)
                print("      [js hint] '%s' found; context: %s" % (kw, txt[max(0, i - 90):i + 210].replace("\n", " ")))

    print()
    print("=" * 72)
    print("STEP 3 - LOOK INSIDE ONE WORKBOOK")
    print("=" * 72)

    if not candidates:
        print("  No direct .xls/.xlsx links found on the page.")
        print("  This usually means the list is built by JavaScript from an API.")
        print("  Check the [js hint] lines above for the endpoint it calls.")
        return

    target = candidates[0]
    print("  Downloading: %s" % target)
    try:
        status, ctype, blob = get(target, timeout=90)
        print("  Got %d bytes  [%s]" % (len(blob), ctype))
    except Exception as e:
        print("  Download failed: %s: %s" % (type(e).__name__, e))
        return

    try:
        import openpyxl
    except ImportError:
        print("  openpyxl not installed - add it to the workflow.")
        return

    try:
        wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    except Exception as e:
        print("  Not a readable .xlsx (%s). It may be legacy .xls, which needs xlrd." % e)
        print("  First 200 bytes:", blob[:200])
        return

    print("  Sheets (%d): %s" % (len(wb.sheetnames), wb.sheetnames[:25]))
    sheet = wb[wb.sheetnames[0]]
    print("\n  --- first 25 rows of '%s' ---" % wb.sheetnames[0])
    for i, row in enumerate(sheet.iter_rows(max_row=25, max_col=10, values_only=True)):
        cells = ["" if c is None else str(c)[:26] for c in row]
        if any(cells):
            print("   %2d | %s" % (i + 1, " | ".join(cells)))
    print("\n  These column headers are what I need to write the parser.")


if __name__ == "__main__":
    main()
